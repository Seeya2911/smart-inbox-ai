"""Unit tests for ml.xlsx_utils — the shared XLSX/XML sanitization utilities.

These tests are independent of any specific exporter and test the core guarantees
of the sanitization and validation layer.
"""
from __future__ import annotations

import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from ml.xlsx_utils import (
    EXCEL_MAX_CELL_CHARS,
    TRUNCATION_MARKER,
    prepare_cell,
    sanitize_for_xlsx,
    truncate_for_xlsx,
    validate_workbook,
)


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _make_workbook(path: Path, rows: list[tuple]) -> None:
    """Write a minimal 'emails'-sheet workbook for validation tests."""
    wb = Workbook()
    ws = wb.active
    ws.title = "emails"
    ws.append(["subject", "body", "source"])
    for row in rows:
        ws.append(list(row))
    wb.save(path)


# ---------------------------------------------------------------------------
# sanitize_for_xlsx
# ---------------------------------------------------------------------------

class TestSanitizeForXlsx:
    """Low-level character-removal tests."""

    # -- Illegal chars -------------------------------------------------------

    def test_nul_removed(self):
        assert sanitize_for_xlsx("a\x00b") == "ab"

    def test_c0_range_removed(self):
        """All C0 controls except \\x09 \\x0A \\x0D must be removed."""
        for cp in range(0x00, 0x09):
            assert sanitize_for_xlsx(chr(cp)) == "", f"chr({cp:#x}) should be removed"
        for cp in range(0x0B, 0x0D):
            assert sanitize_for_xlsx(chr(cp)) == "", f"chr({cp:#x}) should be removed"
        for cp in range(0x0E, 0x20):
            assert sanitize_for_xlsx(chr(cp)) == "", f"chr({cp:#x}) should be removed"

    def test_del_removed(self):
        assert sanitize_for_xlsx("\x7f") == ""

    def test_c1_range_removed(self):
        """All C1 controls (0x80–0x9F) are illegal in XML 1.0."""
        for cp in range(0x80, 0xA0):
            assert sanitize_for_xlsx(chr(cp)) == "", f"chr({cp:#x}) should be removed"

    def test_surrogates_removed(self):
        """Lone surrogates are not valid Unicode scalar values."""
        assert sanitize_for_xlsx("\udcff") == ""
        assert sanitize_for_xlsx("\ud800") == ""

    def test_fffe_removed(self):
        assert sanitize_for_xlsx("\uFFFE") == ""

    def test_ffff_removed(self):
        assert sanitize_for_xlsx("\uFFFF") == ""

    # -- Legal chars ---------------------------------------------------------

    def test_tab_preserved(self):
        assert sanitize_for_xlsx("\x09") == "\x09"

    def test_lf_preserved(self):
        assert sanitize_for_xlsx("\x0A") == "\x0A"

    def test_cr_preserved(self):
        assert sanitize_for_xlsx("\x0D") == "\x0D"

    def test_printable_ascii_preserved(self):
        text = "".join(chr(c) for c in range(0x20, 0x7F))
        assert sanitize_for_xlsx(text) == text

    def test_latin1_supplement_legal_chars_preserved(self):
        """0xA0–0xFF (Latin-1 supplement) are legal in XML 1.0."""
        text = "".join(chr(c) for c in range(0xA0, 0x100))
        assert sanitize_for_xlsx(text) == text

    def test_high_unicode_preserved(self):
        """Supplementary plane chars (emoji, CJK extensions) are legal in XML."""
        text = "😀🎉🔥"
        assert sanitize_for_xlsx(text) == text

    def test_empty_string_unchanged(self):
        assert sanitize_for_xlsx("") == ""

    def test_clean_email_unchanged(self):
        text = "Dear Alice,\n\nPlease review the attached report.\n\nBest,\nBob"
        assert sanitize_for_xlsx(text) == text


# ---------------------------------------------------------------------------
# truncate_for_xlsx
# ---------------------------------------------------------------------------

class TestTruncateForXlsx:
    def test_short_not_truncated(self):
        text = "hello"
        result, was = truncate_for_xlsx(text)
        assert result == text
        assert not was

    def test_exactly_at_limit_not_truncated(self):
        text = "a" * EXCEL_MAX_CELL_CHARS
        result, was = truncate_for_xlsx(text)
        assert result == text
        assert not was

    def test_one_over_limit_truncated(self):
        text = "a" * (EXCEL_MAX_CELL_CHARS + 1)
        result, was = truncate_for_xlsx(text)
        assert was
        assert len(result) == EXCEL_MAX_CELL_CHARS
        assert result.endswith(TRUNCATION_MARKER)

    def test_very_long_truncated(self):
        text = "x" * 200_000
        result, was = truncate_for_xlsx(text)
        assert was
        assert len(result) == EXCEL_MAX_CELL_CHARS

    def test_custom_max(self):
        result, was = truncate_for_xlsx("abcdefgh", max_chars=5)
        assert was
        assert len(result) == 5

    def test_custom_marker(self):
        result, was = truncate_for_xlsx("abcdefgh", max_chars=6, marker="…")
        assert was
        assert result.endswith("…")
        assert len(result) == 6


# ---------------------------------------------------------------------------
# prepare_cell (combined)
# ---------------------------------------------------------------------------

class TestPrepareCell:
    def test_sanitizes_then_truncates(self):
        # 100_000 x's with some illegal chars sprinkled in
        text = "x" * 50_000 + "\x00" * 10 + "x" * 50_000
        result, was_truncated = prepare_cell(text)
        # No illegal chars
        assert "\x00" not in result
        # Length constraint respected
        assert len(result) <= EXCEL_MAX_CELL_CHARS
        assert was_truncated

    def test_clean_short_text_unchanged(self):
        text = "Hello, World!"
        result, was = prepare_cell(text)
        assert result == text
        assert not was


# ---------------------------------------------------------------------------
# validate_workbook
# ---------------------------------------------------------------------------

class TestValidateWorkbook:
    def test_valid_workbook_passes(self, tmp_path: Path):
        path = tmp_path / "ok.xlsx"
        _make_workbook(path, [
            ("Hello", "World body", "enron"),
            ("", "Cooper, did you add security?", "enron"),
            ("Win", "Congratulations!", "spam_corpus"),
        ])
        result = validate_workbook(path)
        assert result["valid"], result["errors"]
        assert result["row_count"] == 3
        assert result["errors"] == []

    def test_missing_file_fails(self, tmp_path: Path):
        result = validate_workbook(tmp_path / "nonexistent.xlsx")
        assert not result["valid"]
        assert any("does not exist" in e for e in result["errors"])

    def test_wrong_headers_fail(self, tmp_path: Path):
        path = tmp_path / "bad_headers.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "emails"
        ws.append(["subject", "body", "WRONG_COLUMN"])
        ws.append(["a", "b", "enron"])
        wb.save(path)
        result = validate_workbook(path)
        assert not result["valid"]
        assert any("Headers mismatch" in e for e in result["errors"])

    def test_missing_emails_sheet_fails(self, tmp_path: Path):
        path = tmp_path / "no_emails.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"  # wrong name
        ws.append(["subject", "body", "source"])
        ws.append(["a", "b", "enron"])
        wb.save(path)
        result = validate_workbook(path)
        assert not result["valid"]
        assert any("emails" in e for e in result["errors"])

    def test_empty_sheet_fails(self, tmp_path: Path):
        path = tmp_path / "empty.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "emails"
        # No rows at all
        wb.save(path)
        result = validate_workbook(path)
        assert not result["valid"]

    def test_header_only_fails(self, tmp_path: Path):
        path = tmp_path / "header_only.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "emails"
        ws.append(["subject", "body", "source"])
        # No data rows
        wb.save(path)
        result = validate_workbook(path)
        assert not result["valid"]
        assert any("no data rows" in e.lower() for e in result["errors"])

    def test_clean_workbook_validates_successfully(self, tmp_path: Path):
        """A workbook with only clean content must pass validate_workbook()."""
        path = tmp_path / "clean.xlsx"
        _make_workbook(path, [("Normal", "Clean body", "enron")])
        result = validate_workbook(path)
        assert result["valid"]
        assert result["errors"] == []

    def test_source_counts_correct(self, tmp_path: Path):
        path = tmp_path / "multi_source.xlsx"
        _make_workbook(path, [
            ("A", "Body A", "enron"),
            ("B", "Body B", "enron"),
            ("C", "Body C", "spam_corpus"),
            ("D", "Body D", "phishing_corpus"),
        ])
        result = validate_workbook(path)
        assert result["valid"]
        assert result["source_counts"]["enron"] == 2
        assert result["source_counts"]["spam_corpus"] == 1
        assert result["source_counts"]["phishing_corpus"] == 1

    def test_row_count_correct(self, tmp_path: Path):
        path = tmp_path / "rows.xlsx"
        _make_workbook(path, [
            ("A", "B", "enron"),
            ("C", "D", "spam_corpus"),
            ("E", "F", "phishing_corpus"),
        ])
        result = validate_workbook(path)
        assert result["row_count"] == 3

    def test_non_xlsx_file_fails(self, tmp_path: Path):
        """A plain text file masquerading as XLSX must fail validation."""
        path = tmp_path / "fake.xlsx"
        path.write_bytes(b"this is not a zip file")
        result = validate_workbook(path)
        assert not result["valid"]
        assert any("ZIP" in e or "zip" in e for e in result["errors"])

    def test_truncation_marker_counted(self, tmp_path: Path):
        """validate_workbook must report the count of truncated cells."""
        path = tmp_path / "truncated.xlsx"
        _make_workbook(path, [("Subject", "Body" + TRUNCATION_MARKER, "enron")])
        result = validate_workbook(path)
        assert result["valid"]
        assert result["truncated_cells"] >= 1

    def test_multiline_body_survives(self, tmp_path: Path):
        """Multi-line email bodies must survive a workbook round-trip."""
        path = tmp_path / "multiline.xlsx"
        body = "Line 1\nLine 2\nLine 3"
        _make_workbook(path, [("Hello", body, "enron")])
        result = validate_workbook(path)
        assert result["valid"]
        # Re-read and verify the newlines survived
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        values = list(wb["emails"].iter_rows(values_only=True))
        assert values[1][1] == body

    def test_unicode_survives_roundtrip(self, tmp_path: Path):
        """Accented characters and CJK must survive the XLSX round-trip."""
        path = tmp_path / "unicode.xlsx"
        body = "café résumé 電子 メール 🎉"
        _make_workbook(path, [("Unicode test", body, "enron")])
        result = validate_workbook(path)
        assert result["valid"]
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        values = list(wb["emails"].iter_rows(values_only=True))
        assert values[1][1] == body
