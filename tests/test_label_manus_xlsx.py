import json
from pathlib import Path

from openpyxl import Workbook

from ml.label_manus_xlsx import label_workbook, load_workbook_examples


class FakeLLM:
    model = "fake"

    def classify(self, subject, body):
        return {
            "intent": "security" if "password" in body.lower() else "question",
            "priority": "high" if "password" in body.lower() else "medium",
            "priority_reasons": ["security impact"] if "password" in body.lower() else ["response requested"],
            "intent_reason": "security event" if "password" in body.lower() else "asks for information",
            "priority_reason": "prompt action is appropriate" if "password" in body.lower() else "timely response is useful",
            "confidence": 0.93,
        }


def _write_fixture(tmp_path: Path) -> tuple[Path, Path]:
    xlsx = tmp_path / "corpus.xlsx"
    manifest = tmp_path / "corpus.manifest.json"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["subject", "body", "source"])
    sheet.append(["Account alert", "Your password was changed. If this was not you, act now.", "enron"])
    sheet.append(["Meeting", "What time should we meet tomorrow?", "spam_corpus"])
    workbook.save(xlsx)

    manifest.write_text(
        json.dumps(
            {
                "output_rows": 2,
                "row_map": [
                    {"output_row": 2, "id": "enron_1", "source_example_id": "1", "source": "enron", "language": "en", "provenance": "fixture"},
                    {"output_row": 3, "id": "spam_2", "source_example_id": "2", "source": "spam_corpus", "language": "en", "provenance": "fixture"},
                ],
            }
        ),
        encoding="utf-8",
    )
    return xlsx, manifest


def test_load_workbook_examples_joins_manifest(tmp_path):
    xlsx, manifest = _write_fixture(tmp_path)
    examples = load_workbook_examples(xlsx, manifest)

    assert [example.id for example in examples] == ["enron_1", "spam_2"]
    assert examples[0].source_example_id == "1"
    assert examples[1].source == "spam_corpus"


def test_llm_is_authoritative_and_output_is_resumable(tmp_path):
    xlsx, manifest = _write_fixture(tmp_path)
    output = tmp_path / "labeled.jsonl"

    first = label_workbook(xlsx, manifest, output, llm_client=FakeLLM())
    assert first["newly_labeled_rows"] == 2
    assert first["llm_rule_disagreements_new"] >= 0

    records = [json.loads(line) for line in output.read_text(encoding="utf-8").splitlines()]
    assert records[0]["intent"] == "security"
    assert records[0]["priority"] == "high"
    assert records[0]["label_source"] == "llm"
    assert records[0]["source_example_id"] == "1"

    second = label_workbook(xlsx, manifest, output, llm_client=FakeLLM(), resume=True)
    assert second["newly_labeled_rows"] == 0
    assert len(output.read_text(encoding="utf-8").splitlines()) == 2
