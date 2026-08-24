"""Regression tests for ml.export_manus_dataset.

Covers all failure modes described in the fix/manus-xlsx-validation issue:
- Illegal control characters (C0, DEL, C1, surrogates)
- Normal Unicode preservation
- Newline / tab preservation
- Excel cell-length truncation
- Missing / present subject fields
- Correct column order
- Workbook can be re-opened after export
- Workbook ZIP/XML structure is valid
- Multiple corpus sources
- Deduplication behaviour
- Enron subject extraction path
"""
from __future__ import annotations

import json
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from ml.export_manus_dataset import export_xlsx, load_rows, write_manifest
from ml.xlsx_utils import (
    EXCEL_MAX_CELL_CHARS,
    TRUNCATION_MARKER,
    prepare_cell,
    sanitize_for_xlsx,
    validate_workbook,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_jsonl(path: Path, rows: list[dict]) -> None:
    path.write_text(
        "".join(json.dumps(r, ensure_ascii=False) + "\n" for r in rows),
        encoding="utf-8",
    )


def _xlsx_values(path: Path) -> list[tuple]:
    """Return all rows (including header) from the 'emails' sheet."""
    wb = load_workbook(path, read_only=True)
    return list(wb["emails"].iter_rows(values_only=True))


# ---------------------------------------------------------------------------
# sanitize_for_xlsx unit tests
# ---------------------------------------------------------------------------

class TestSanitizeForXlsx:
    """Unit tests for the XML 1.0 illegal-character filter."""

    def test_c0_controls_removed(self):
        """NUL through BEL, BS, and other C0 controls are removed."""
        text = "hello\x00world\x01\x02\x03\x04\x05\x06\x07\x08"
        assert sanitize_for_xlsx(text) == "helloworld"

    def test_vertical_tab_and_form_feed_removed(self):
        assert sanitize_for_xlsx("a\x0bb\x0cc") == "abc"

    def test_c0_tail_controls_removed(self):
        """0x0E–0x1F are removed."""
        text = "".join(chr(c) for c in range(0x0E, 0x20))
        assert sanitize_for_xlsx(text) == ""

    def test_del_character_removed(self):
        """\\x7F (DEL) is illegal in XML 1.0 and must be removed."""
        assert sanitize_for_xlsx("abc\x7fdef") == "abcdef"

    def test_c1_controls_removed(self):
        """\\x80–\\x9F (C1 controls) are illegal in XML 1.0."""
        text = "".join(chr(c) for c in range(0x80, 0xA0))
        assert sanitize_for_xlsx(text) == ""

    def test_unicode_surrogates_removed(self):
        """Lone surrogates (\\uD800–\\uDFFF) are not valid Unicode scalars."""
        # Python allows lone surrogates via surrogatepass; they are illegal in XML.
        text = "before\udcffafter"
        assert sanitize_for_xlsx(text) == "beforeafter"

    def test_fffe_and_ffff_removed(self):
        """\\uFFFE and \\uFFFF are permanently unassigned and illegal in XML 1.0."""
        assert sanitize_for_xlsx("a\uFFFEb\uFFFFc") == "abc"

    def test_tab_preserved(self):
        """Tab (\\x09) is legal in XML 1.0 and useful in email bodies."""
        assert sanitize_for_xlsx("col1\tcol2") == "col1\tcol2"

    def test_lf_preserved(self):
        """Line Feed (\\x0A) must survive — Excel supports multi-line cells."""
        assert sanitize_for_xlsx("line1\nline2") == "line1\nline2"

    def test_cr_normalized_to_lf(self):
        """Carriage Return (\\x0D) is normalized to standard Line Feed (\\n)."""
        assert sanitize_for_xlsx("a\rb") == "a\nb"
        assert sanitize_for_xlsx("a\r\nb") == "a\nb"

    def test_normal_ascii_preserved(self):
        text = "Hello, World! 123 #@$%^&*()"
        assert sanitize_for_xlsx(text) == text

    def test_accented_characters_preserved(self):
        text = "café résumé naïve"
        assert sanitize_for_xlsx(text) == text

    def test_cjk_characters_preserved(self):
        text = "电子邮件 メール 이메일"
        assert sanitize_for_xlsx(text) == text

    def test_emoji_preserved(self):
        """Emoji are above U+FFFF and are legal in XML."""
        text = "Meeting 📅 tomorrow 🎉"
        assert sanitize_for_xlsx(text) == text

    def test_enron_style_mixed_content_cleaned(self):
        """Reproduces the style of control chars seen in Enron exports."""
        raw = "Enron\x01s value add\x08 update"
        assert sanitize_for_xlsx(raw) == "Enrons value add update"

    def test_empty_string_unchanged(self):
        assert sanitize_for_xlsx("") == ""

    def test_none_like_empty_handled_by_prepare_cell(self):
        """prepare_cell wraps sanitize; empty string is a no-op."""
        result, truncated = prepare_cell("")
        assert result == ""
        assert not truncated


# ---------------------------------------------------------------------------
# Truncation tests
# ---------------------------------------------------------------------------

class TestTruncationForXlsx:
    def test_short_text_not_truncated(self):
        text = "a" * 100
        result, truncated = prepare_cell(text)
        assert result == text
        assert not truncated

    def test_exactly_at_limit_not_truncated(self):
        text = "a" * EXCEL_MAX_CELL_CHARS
        result, truncated = prepare_cell(text)
        assert result == text
        assert not truncated

    def test_over_limit_truncated_with_marker(self):
        text = "a" * (EXCEL_MAX_CELL_CHARS + 500)
        result, truncated = prepare_cell(text)
        assert truncated
        assert len(result) == EXCEL_MAX_CELL_CHARS
        assert result.endswith(TRUNCATION_MARKER)

    def test_truncated_result_fits_in_cell(self):
        """Truncated string must be exactly EXCEL_MAX_CELL_CHARS long."""
        long_body = "x" * 100_000
        result, truncated = prepare_cell(long_body)
        assert truncated
        assert len(result) <= EXCEL_MAX_CELL_CHARS


# ---------------------------------------------------------------------------
# load_rows tests
# ---------------------------------------------------------------------------

class TestLoadRows:
    def test_removes_empty_rows(self, tmp_path: Path):
        source = tmp_path / "raw.jsonl"
        _make_jsonl(source, [
            {"id": "1", "subject": "Hello", "body": "World", "source": "enron"},
            {"id": "2", "subject": "", "body": "", "source": "enron"},
        ])
        rows, stats = load_rows([source])
        assert len(rows) == 1
        assert stats["empty_removed"] == 1

    def test_removes_exact_duplicates(self, tmp_path: Path):
        source = tmp_path / "raw.jsonl"
        _make_jsonl(source, [
            {"id": "1", "subject": "Hello", "body": "World", "source": "enron"},
            {"id": "2", "subject": " hello ", "body": "world", "source": "spam_corpus"},
            {"id": "3", "subject": "Different", "body": "Message", "source": "phishing_corpus"},
        ])
        rows, stats = load_rows([source])
        assert len(rows) == 2
        assert stats["exact_duplicates_removed"] == 1

    def test_multiple_sources_preserved(self, tmp_path: Path):
        source = tmp_path / "raw.jsonl"
        _make_jsonl(source, [
            {"subject": "A", "body": "B", "source": "enron"},
            {"subject": "C", "body": "D", "source": "spam_corpus"},
            {"subject": "E", "body": "F", "source": "phishing_corpus"},
        ])
        rows, stats = load_rows([source])
        sources = {r["source"] for r in rows}
        assert sources == {"enron", "spam_corpus", "phishing_corpus"}

    def test_raises_when_source_missing(self, tmp_path: Path):
        source = tmp_path / "raw.jsonl"
        _make_jsonl(source, [{"subject": "Hi", "body": "Body", "source": ""}])
        with pytest.raises(ValueError, match="non-empty source"):
            load_rows([source])

    def test_subject_genuinely_absent_is_allowed(self, tmp_path: Path):
        """Emails with no subject but a body must pass through (enron pattern)."""
        source = tmp_path / "raw.jsonl"
        _make_jsonl(source, [
            {"id": "e1", "subject": "", "body": "Cooper, did you add security?", "source": "enron"},
        ])
        rows, stats = load_rows([source])
        assert len(rows) == 1
        assert rows[0]["subject"] == ""
        assert "Cooper" in rows[0]["body"]


# ---------------------------------------------------------------------------
# export_xlsx tests
# ---------------------------------------------------------------------------

class TestExportXlsx:
    def test_column_order_is_subject_body_source(self, tmp_path: Path):
        output = tmp_path / "manus.xlsx"
        export_xlsx(
            [{"subject": "A", "body": "B", "source": "enron"}],
            output,
        )
        values = _xlsx_values(output)
        assert values[0] == ("subject", "body", "source")
        assert values[1] == ("A", "B", "enron")

    def test_illegal_c0_control_chars_removed(self, tmp_path: Path):
        output = tmp_path / "manus.xlsx"
        export_xlsx(
            [{"subject": "Sub\x00ject", "body": "Body\x01text", "source": "enron"}],
            output,
        )
        values = _xlsx_values(output)
        assert values[1][0] == "Subject"   # NUL removed → "Subject"
        assert values[1][1] == "Bodytext"  # SOH removed

    def test_del_character_removed(self, tmp_path: Path):
        """\\x7F must be removed (was missing from the old regex)."""
        output = tmp_path / "manus.xlsx"
        export_xlsx(
            [{"subject": "abc\x7fdef", "body": "normal", "source": "enron"}],
            output,
        )
        values = _xlsx_values(output)
        assert values[1][0] == "abcdef"

    def test_c1_control_chars_removed(self, tmp_path: Path):
        """C1 controls (0x80–0x9F) are illegal in XML 1.0."""
        output = tmp_path / "manus.xlsx"
        dirty = "".join(chr(c) for c in range(0x80, 0xA0))
        export_xlsx(
            [{"subject": dirty + "clean", "body": "ok", "source": "enron"}],
            output,
        )
        values = _xlsx_values(output)
        assert values[1][0] == "clean"

    def test_newlines_preserved(self, tmp_path: Path):
        """Multi-line email bodies must survive the XLSX round-trip."""
        output = tmp_path / "manus.xlsx"
        export_xlsx(
            [{"subject": "Hello\nWorld", "body": "Line1\nLine2\nLine3", "source": "enron"}],
            output,
        )
        values = _xlsx_values(output)
        assert values[1][0] == "Hello\nWorld"
        assert values[1][1] == "Line1\nLine2\nLine3"

    def test_tabs_preserved(self, tmp_path: Path):
        """Tabs are legal in XML 1.0 and may appear in formatted email bodies."""
        output = tmp_path / "manus.xlsx"
        export_xlsx(
            [{"subject": "Tab\there", "body": "Col1\tCol2", "source": "enron"}],
            output,
        )
        values = _xlsx_values(output)
        assert values[1][0] == "Tab\there"
        assert values[1][1] == "Col1\tCol2"

    def test_unicode_preserved(self, tmp_path: Path):
        """Accented characters, CJK, and emoji must survive the round-trip."""
        output = tmp_path / "manus.xlsx"
        export_xlsx(
            [{"subject": "café", "body": "メール 🎉", "source": "enron"}],
            output,
        )
        values = _xlsx_values(output)
        assert values[1][0] == "café"
        assert values[1][1] == "メール 🎉"

    def test_missing_subject_handled(self, tmp_path: Path):
        """Empty subject is preserved as empty/None in openpyxl."""
        output = tmp_path / "manus.xlsx"
        export_xlsx(
            [{"subject": "", "body": "Cooper, did you add security?", "source": "enron"}],
            output,
        )
        values = _xlsx_values(output)
        assert values[1][0] in ("", None)
        assert "Cooper" in values[1][1]

    def test_long_body_truncated_at_excel_limit(self, tmp_path: Path):
        """Bodies exceeding 32 767 characters must be truncated with a marker."""
        output = tmp_path / "manus.xlsx"
        long_body = "x" * 100_000
        stats = export_xlsx(
            [{"subject": "Long", "body": long_body, "source": "enron"}],
            output,
        )
        assert stats["truncated_rows"] == 1
        values = _xlsx_values(output)
        body_cell = values[1][1]
        assert len(body_cell) == EXCEL_MAX_CELL_CHARS
        assert body_cell.endswith(TRUNCATION_MARKER)

    def test_normal_body_not_truncated(self, tmp_path: Path):
        output = tmp_path / "manus.xlsx"
        normal_body = "Normal email body."
        stats = export_xlsx(
            [{"subject": "Normal", "body": normal_body, "source": "enron"}],
            output,
        )
        assert stats["truncated_rows"] == 0
        values = _xlsx_values(output)
        assert values[1][1] == normal_body

    def test_workbook_can_be_reopened(self, tmp_path: Path):
        """The generated XLSX must be re-openable without exception."""
        output = tmp_path / "manus.xlsx"
        export_xlsx(
            [{"subject": "A", "body": "B", "source": "enron"}],
            output,
        )
        # Must not raise
        wb = load_workbook(output)
        assert "emails" in wb.sheetnames

    def test_workbook_is_valid_zip(self, tmp_path: Path):
        """XLSX is a ZIP archive; it must be readable as such."""
        output = tmp_path / "manus.xlsx"
        export_xlsx(
            [{"subject": "A", "body": "B", "source": "enron"}],
            output,
        )
        assert zipfile.is_zipfile(output)
        with zipfile.ZipFile(output, "r") as zf:
            assert "xl/worksheets/sheet1.xml" in zf.namelist()

    def test_workbook_xml_is_valid_utf8(self, tmp_path: Path):
        """The sheet XML must be valid UTF-8 (encoding corruption check)."""
        output = tmp_path / "manus.xlsx"
        export_xlsx(
            [{"subject": "Hello", "body": "World", "source": "enron"}],
            output,
        )
        with zipfile.ZipFile(output, "r") as zf:
            raw = zf.read("xl/worksheets/sheet1.xml")
        raw.decode("utf-8")  # must not raise

    def test_multiple_sources_all_present(self, tmp_path: Path):
        """All three supported corpus sources must appear in the output."""
        output = tmp_path / "manus.xlsx"
        export_xlsx(
            [
                {"subject": "A", "body": "Enron msg", "source": "enron"},
                {"subject": "B", "body": "Spam msg", "source": "spam_corpus"},
                {"subject": "C", "body": "Phish msg", "source": "phishing_corpus"},
            ],
            output,
        )
        values = _xlsx_values(output)
        sources = {row[2] for row in values[1:]}
        assert sources == {"enron", "spam_corpus", "phishing_corpus"}

    def test_no_extra_columns(self, tmp_path: Path):
        """The workbook must have exactly three columns (no accidental extras)."""
        output = tmp_path / "manus.xlsx"
        export_xlsx(
            [{"subject": "A", "body": "B", "source": "enron", "intent": "request"}],
            output,
        )
        values = _xlsx_values(output)
        for row in values:
            assert len(row) == 3, f"Expected 3 columns, got {len(row)}: {row}"

    def test_validate_workbook_passes_on_clean_output(self, tmp_path: Path):
        """validate_workbook() must report valid=True for a freshly exported file."""
        output = tmp_path / "manus.xlsx"
        export_xlsx(
            [
                {"subject": "Meeting", "body": "Can we meet at 3 PM?", "source": "enron"},
                {"subject": "Win", "body": "Congratulations!", "source": "spam_corpus"},
            ],
            output,
        )
        result = validate_workbook(output)
        assert result["valid"], f"Validation errors: {result['errors']}"
        assert result["row_count"] == 2
        assert result["source_counts"].get("enron") == 1
        assert result["source_counts"].get("spam_corpus") == 1

    def test_formula_prefixed_strings_saved_as_text_not_formula(self, tmp_path: Path):
        """Strings starting with '=', '===', or MIME headers must not create <f> tags."""
        output = tmp_path / "formula_test.xlsx"
        export_xlsx(
            [
                {
                    "subject": "=?UTF-8?B?4pyoIEhlbGxv?=",
                    "body": "================== Original Message ==================\n=20\n=3D",
                    "source": "enron",
                },
                {
                    "subject": "=HYPERLINK(\"http://phish.com\")",
                    "body": "+1-800-555-0199 / -50% discount / @everyone",
                    "source": "phishing_corpus",
                },
            ],
            output,
        )
        result = validate_workbook(output)
        assert result["valid"], f"Validation failed on formula-prefixed text: {result['errors']}"
        
        # Verify no <f> formula tags exist in the sheet XML
        with zipfile.ZipFile(output, "r") as zf:
            sheet_xml = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
            assert "<f>" not in sheet_xml and "<f " not in sheet_xml, "Formula tags generated in sheet XML!"
        
        # Verify values read back properly
        values = _xlsx_values(output)
        assert values[1][0] == "=?UTF-8?B?4pyoIEhlbGxv?="
        assert "Original Message" in values[1][1]


# ---------------------------------------------------------------------------
# Enron subject extraction regression
# ---------------------------------------------------------------------------

class TestEnronSubjectExtraction:
    """Verify that Enron subject fields are handled correctly end-to-end.

    The corbt/enron-emails Parquet has a 'subject' column distinct from 'body'.
    Subjects may be genuinely absent in internal messages — this is not a bug.
    """

    def test_enron_subject_from_row_field(self, tmp_path: Path):
        """When a 'subject' key is present in the JSONL record, it must be used."""
        source = tmp_path / "enron.jsonl"
        _make_jsonl(source, [
            {
                "id": "enron_42",
                "subject": "Project update",
                "body": "The work is progressing.",
                "source": "enron",
            }
        ])
        rows, _ = load_rows([source])
        assert rows[0]["subject"] == "Project update"

    def test_enron_blank_subject_is_not_a_bug(self, tmp_path: Path):
        """Internal Enron emails sometimes have no subject — preserve as empty."""
        source = tmp_path / "enron.jsonl"
        _make_jsonl(source, [
            {
                "id": "enron_99",
                "subject": "",
                "body": "Cooper, did you add some more security to the expost hourly summary?",
                "source": "enron",
            }
        ])
        rows, _ = load_rows([source])
        assert rows[0]["subject"] == ""
        assert rows[0]["body"].startswith("Cooper")

    def test_enron_subject_survives_xlsx_export(self, tmp_path: Path):
        """Enron subject must appear in column A of the exported workbook."""
        output = tmp_path / "manus.xlsx"
        export_xlsx(
            [{"subject": "Project update", "body": "Progressing.", "source": "enron"}],
            output,
        )
        values = _xlsx_values(output)
        assert values[1][0] == "Project update"
        assert values[1][2] == "enron"


# ---------------------------------------------------------------------------
# write_manifest
# ---------------------------------------------------------------------------

class TestWriteManifest:
    def test_manifest_records_truncation_count(self, tmp_path: Path):
        manifest_path = tmp_path / "manifest.json"
        rows = [{"subject": "A", "body": "B", "source": "enron"}]
        stats = {"input_rows": 1, "empty_removed": 0, "exact_duplicates_removed": 0,
                 "truncated_rows": 0}
        write_manifest(rows, stats, manifest_path)
        payload = json.loads(manifest_path.read_text(encoding="utf-8"))
        assert "truncated_rows" in payload
        assert payload["source_counts"] == {"enron": 1}
        assert payload["output_rows"] == 1
        assert "excel_max_cell_chars" in payload
