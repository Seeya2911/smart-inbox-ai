import json
from pathlib import Path

from openpyxl import load_workbook

from ml.prepare_manus_dataset import prepare_records, write_batches


def _write_jsonl(path: Path, rows: list[dict]) -> None:
    path.write_text("\n".join(json.dumps(row) for row in rows) + "\n", encoding="utf-8")


def test_prepare_records_removes_empty_and_exact_duplicates(tmp_path: Path) -> None:
    source = tmp_path / "raw.jsonl"
    _write_jsonl(
        source,
        [
            {"id": "enron_1", "subject": "Hello", "body": "Discuss budget", "source": "enron"},
            {"id": "spam_1", "subject": "Hello", "body": "Discuss budget", "source": "spam_corpus"},
            {"id": "enron_2", "subject": "", "body": "", "source": "enron"},
        ],
    )

    records, stats = prepare_records([source])

    assert len(records) == 1
    assert records[0]["id"] == "enron_1"
    assert stats == {"rows_read": 3, "empty_removed": 1, "exact_duplicates_removed": 1}


def test_write_batches_contains_only_manus_input_columns(tmp_path: Path) -> None:
    records = [
        {"id": "enron_1", "subject": "A", "body": "Body A", "source": "enron"},
        {"id": "spam_1", "subject": "B", "body": "Body B", "source": "spam_corpus"},
        {"id": "phishing_1", "subject": "C", "body": "Body C", "source": "phishing_corpus"},
    ]

    manifest = write_batches(records, tmp_path, batch_size=2)

    first = load_workbook(tmp_path / "manus_batch_0001.xlsx", read_only=True).active
    assert list(first.values) == [
        ("subject", "body", "source"),
        ("A", "Body A", "enron"),
        ("B", "Body B", "spam_corpus"),
    ]
    assert len(manifest) == 3
    assert manifest[0] == {"batch": "manus_batch_0001.xlsx", "excel_row": 2, "id": "enron_1", "source": "enron"}
    assert manifest[-1]["batch"] == "manus_batch_0002.xlsx"


def test_write_batches_sanitizes_illegal_control_characters(tmp_path: Path) -> None:
    """write_batches() must strip C0/DEL/C1/surrogate chars before writing cells."""
    records = [
        {
            "id": "enron_1",
            "subject": "Sub\x00ject",    # NUL must be removed
            "body": "Body\x7ftext\x80",  # DEL and C1 must be removed
            "source": "enron",
        }
    ]
    write_batches(records, tmp_path, batch_size=10)
    ws = load_workbook(tmp_path / "manus_batch_0001.xlsx", read_only=True).active
    values = list(ws.iter_rows(values_only=True))
    subject_cell = values[1][0]
    body_cell = values[1][1]
    assert "\x00" not in subject_cell, "NUL must be removed from subject"
    assert "\x7f" not in body_cell, "DEL must be removed from body"
    assert "\x80" not in body_cell, "C1 control must be removed from body"


def test_write_batches_truncates_long_bodies(tmp_path: Path) -> None:
    """Bodies exceeding Excel's 32 767-character limit must be truncated."""
    from ml.xlsx_utils import EXCEL_MAX_CELL_CHARS, TRUNCATION_MARKER

    records = [
        {
            "id": "enron_1",
            "subject": "Long email",
            "body": "x" * 100_000,
            "source": "enron",
        }
    ]
    write_batches(records, tmp_path, batch_size=10)
    ws = load_workbook(tmp_path / "manus_batch_0001.xlsx", read_only=True).active
    values = list(ws.iter_rows(values_only=True))
    body_cell = values[1][1]
    assert len(body_cell) == EXCEL_MAX_CELL_CHARS
    assert body_cell.endswith(TRUNCATION_MARKER)
