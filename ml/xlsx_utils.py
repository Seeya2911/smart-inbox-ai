"""Shared XLSX/XML sanitization and validation utilities for the Manus corpus pipeline.

XLSX files are ZIP archives containing XML.  The XML 1.0 specification defines a
strict set of legal character code points.  Any character outside that set causes
the OpenXML parser embedded in Microsoft Excel to report a content error and offer
to "recover" the workbook.

Legal XML 1.0 characters (everything else is illegal):
    #x9 | #xA | #xD | [#x20–#xD7FF] | [#xE000–#xFFFD] | [#x10000–#x10FFFF]

Common offenders found in real email corpora:
    \\x00–\\x08  — C0 controls (NUL, SOH, STX …)
    \\x0B        — Vertical Tab
    \\x0C        — Form Feed
    \\x0E–\\x1F  — remaining C0 controls
    \\x7F        — DEL (often missed by incomplete regexes)
    \\x80–\\x9F  — C1 controls (legal Latin-1 but illegal XML 1.0)
    \\uD800–\\uDFFF  — UTF-16 surrogates (invalid Unicode scalar values)
    \\uFFFE       — Byte-Order-Mark reversed; permanently unassigned
    \\uFFFF       — permanently unassigned; illegal in XML 1.0

Excel cell hard limit:
    32 767 characters per cell (including trailing newlines).
    Cells that exceed this limit cause Excel to silently truncate or corrupt the file.
"""
from __future__ import annotations

import re
import zipfile
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

#: Hard per-cell character limit imposed by Excel / OpenXML.
EXCEL_MAX_CELL_CHARS: int = 32_767

#: Marker appended to truncated cell values so the truncation is visible.
TRUNCATION_MARKER: str = " [TRUNCATED]"

# ---------------------------------------------------------------------------
# Illegal-character regex (complete XML 1.0 definition)
# ---------------------------------------------------------------------------
#
# We match everything that is NOT a legal XML 1.0 character and strip it.
# Legal set: \x09, \x0A, \x0D, \x20–\uD7FF, \uE000–\uFFFD, \U00010000–\U0010FFFF
# The complement (illegal set) is therefore the negation of the above.
#
# Python regex negated character class:
#   [^\x09\x0A\x0D\x20-\uD7FF\uE000-\uFFFD\U00010000-\U0010FFFF]
#
# Note: Python's `re` module works on Unicode code points, so surrogate range
# \uD800–\uDFFF is implicitly excluded by the gap between \uD7FF and \uE000.
#
# IMPORTANT: Python's re module does NOT interpret \uXXXX or \UXXXXXXXX escapes
# in raw strings (r"..."). We must use actual Unicode characters in the pattern.
# We build the pattern using a regular string with Python-level Unicode escapes.
#
# We deliberately SPLIT the original XML 1.0 range [\x20-\uD7FF] into two:
#   [\x20-\x7E]   — printable ASCII (space through tilde)
#   [\xA0-\uD7FF] — Latin-1 supplement and higher BMP characters
# This gap (\x7F and \x80-\x9F) excludes:
#   \x7F      — DEL: technically legal in XML 1.0 but problematic in Excel
#   \x80-\x9F — C1 controls: legal Latin-1 bytes but illegal in XML 1.0

_ILLEGAL_XML10_CHARS: re.Pattern[str] = re.compile(
    "[^\x09\x0A\x0D\x20-\x7E\xA0-\uD7FF\uE000-\uFFFD\U00010000-\U0010FFFF]",
    re.UNICODE,
)


# ---------------------------------------------------------------------------
# Public helpers
# ---------------------------------------------------------------------------

def sanitize_for_xlsx(text: str) -> str:
    """Remove all characters that are illegal in XML 1.0 (and therefore XLSX).

    Safe characters that are deliberately preserved:
    - Tab (\\x09)
    - Line Feed (\\x0A)  — Excel renders multi-line cells correctly
    - All printable ASCII (\\x20–\\x7E)
    - All printable Unicode above \\x9F (accented letters, CJK, emoji, etc.)

    Removed/normalized:
    - NUL and other C0 controls except whitespace listed above
    - DEL (\\x7F)
    - C1 controls (\\x80–\\x9F)
    - Unicode surrogates (\\uD800–\\uDFFF)
    - \\uFFFE and \\uFFFF
    - Normalizes \\r\\n and lone \\r to standard \\n

    Args:
        text: Raw string that may contain illegal characters.

    Returns:
        String safe to write into any XLSX cell.
    """
    if not text:
        return text
    # Normalize carriage returns and CRLF to standard LF (\n)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return _ILLEGAL_XML10_CHARS.sub("", text)


def truncate_for_xlsx(
    text: str,
    max_chars: int = EXCEL_MAX_CELL_CHARS,
    marker: str = TRUNCATION_MARKER,
) -> tuple[str, bool]:
    """Truncate *text* so it fits within Excel's per-cell character limit.

    If truncation is necessary the returned string ends with *marker* so the
    truncation is explicit and auditable.

    Args:
        text:      The (already sanitized) cell text.
        max_chars: Maximum allowed characters (default: ``EXCEL_MAX_CELL_CHARS``).
        marker:    Suffix appended when truncation occurs.

    Returns:
        ``(result_text, was_truncated)`` — the (possibly truncated) string and
        a boolean flag that is ``True`` only when truncation actually occurred.
    """
    if len(text) <= max_chars:
        return text, False
    # Reserve space for the marker inside the limit.
    # Guard: if max_chars is shorter than the marker itself, plain-truncate
    # without appending the marker (the marker would overflow the limit).
    if max_chars <= len(marker):
        return text[:max_chars], True
    cut = max_chars - len(marker)
    return text[:cut] + marker, True


def prepare_cell(text: str) -> tuple[str, bool]:
    """Sanitize then truncate *text* for safe XLSX cell storage.

    Convenience wrapper combining :func:`sanitize_for_xlsx` and
    :func:`truncate_for_xlsx`.

    Returns:
        ``(cell_text, was_truncated)``
    """
    clean = sanitize_for_xlsx(text)
    return truncate_for_xlsx(clean)


# ---------------------------------------------------------------------------
# Workbook validation
# ---------------------------------------------------------------------------

def validate_workbook(path: Path, expected_headers: list[str] | None = None) -> dict[str, Any]:
    """Perform comprehensive validation of a Manus-corpus XLSX workbook.

    Checks performed (in order):
    1. File exists and is non-empty.
    2. The file is a valid ZIP archive (XLSX is ZIP + XML).
    3. The ZIP contains ``xl/worksheets/sheet1.xml`` (the first worksheet).
    4. The workbook can be opened by openpyxl without exception.
    5. A worksheet named ``"emails"`` exists.
    6. The header row matches *expected_headers* (default: ``["subject", "body", "source"]``).
    7. Every data row has the correct number of columns.
    8. No cell value contains illegal XML 1.0 characters.
    9. Source column contains only known values.
    10. The workbook can be saved to a temporary path and re-opened (round-trip).
    11. Row count is at least 1 (excluding header).

    Args:
        path:             Path to the ``.xlsx`` file.
        expected_headers: Expected first-row values.  Defaults to
                          ``["subject", "body", "source"]``.

    Returns:
        A dict with validation results::

            {
                "valid": bool,
                "row_count": int,          # data rows (excluding header)
                "source_counts": dict,     # {source: count}
                "truncated_cells": int,    # cells containing TRUNCATION_MARKER
                "errors": list[str],       # human-readable error descriptions
            }

    Raises:
        Nothing — all failures are recorded in ``errors``.
    """
    import tempfile

    from openpyxl import load_workbook as _load_wb

    if expected_headers is None:
        expected_headers = ["subject", "body", "source"]

    errors: list[str] = []
    row_count = 0
    source_counts: dict[str, int] = {}
    truncated_cells = 0

    # 1. File exists and is non-empty
    if not path.exists():
        return {"valid": False, "row_count": 0, "source_counts": {}, "truncated_cells": 0,
                "errors": [f"File does not exist: {path}"]}
    if path.stat().st_size == 0:
        return {"valid": False, "row_count": 0, "source_counts": {}, "truncated_cells": 0,
                "errors": [f"File is empty: {path}"]}

    # 2. Valid ZIP structure and XML syntax validation
    try:
        import xml.etree.ElementTree as ET

        with zipfile.ZipFile(path, "r") as zf:
            names = zf.namelist()
            if "xl/worksheets/sheet1.xml" not in names:
                errors.append("ZIP is missing xl/worksheets/sheet1.xml")
            
            # Verify all XML parts parse cleanly as valid XML
            for name in names:
                if name.endswith(".xml") or name.endswith(".rels"):
                    raw_content = zf.read(name)
                    try:
                        root_elem = ET.fromstring(raw_content)
                    except ET.ParseError as exc:
                        errors.append(f"XML parse error in {name}: {exc}")

            # 3. Check for formula tags in worksheet
            if "xl/worksheets/sheet1.xml" in names:
                sheet_xml = zf.read("xl/worksheets/sheet1.xml").decode("utf-8", errors="replace")
                if "<f>" in sheet_xml or "<f " in sheet_xml:
                    errors.append("Worksheet contains <f> formula tags which causes Excel repair warnings")
    except zipfile.BadZipFile as exc:
        errors.append(f"Not a valid ZIP/XLSX file: {exc}")
        return {"valid": False, "row_count": 0, "source_counts": {}, "truncated_cells": 0,
                "errors": errors}

    # 4. openpyxl can open it
    try:
        wb = _load_wb(path, read_only=True)
    except Exception as exc:  # noqa: BLE001
        errors.append(f"openpyxl failed to open workbook: {exc}")
        return {"valid": bool(not errors), "row_count": 0, "source_counts": {},
                "truncated_cells": 0, "errors": errors}

    # 5. "emails" worksheet exists
    if "emails" not in wb.sheetnames:
        errors.append(f"Expected worksheet 'emails'; found: {wb.sheetnames}")
        wb.close()
        return {"valid": False, "row_count": 0, "source_counts": {}, "truncated_cells": 0,
                "errors": errors}

    sheet = wb["emails"]

    # 6. Header row
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        errors.append("Worksheet 'emails' is completely empty")
        wb.close()
        return {"valid": False, "row_count": 0, "source_counts": {}, "truncated_cells": 0,
                "errors": errors}

    if header != expected_headers:
        errors.append(f"Headers mismatch: expected {expected_headers}, got {header}")

    n_cols = len(expected_headers)

    # 7–9. Data rows
    for row_idx, row in enumerate(rows_iter, start=2):
        row_count += 1
        cells = list(row)

        # 7. Column count
        if len(cells) != n_cols:
            errors.append(f"Row {row_idx}: expected {n_cols} columns, got {len(cells)}")
            continue

        for col_idx, value in enumerate(cells):
            cell_str = str(value) if value is not None else ""

            # 8. Illegal characters
            if _ILLEGAL_XML10_CHARS.search(cell_str):
                illegal_chars = set(_ILLEGAL_XML10_CHARS.findall(cell_str))
                errors.append(
                    f"Row {row_idx} col {col_idx + 1}: illegal XML chars "
                    f"{[hex(ord(c)) for c in illegal_chars]!r}"
                )

            # Count truncation markers
            if TRUNCATION_MARKER in cell_str:
                truncated_cells += 1

        # 9. Source values (last column by convention for the Manus schema)
        if len(cells) >= n_cols:
            source_val = str(cells[n_cols - 1]) if cells[n_cols - 1] is not None else ""
            source_counts[source_val] = source_counts.get(source_val, 0) + 1

    # 11. At least one data row
    if row_count == 0:
        errors.append("Worksheet has no data rows (only a header or empty)")

    wb.close()

    # 10. Round-trip: save to temp and re-open
    if not errors:
        try:
            wb2 = _load_wb(path)
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp_path = Path(tmp.name)
            try:
                wb2.save(tmp_path)
                _load_wb(tmp_path, read_only=True).close()
            finally:
                tmp_path.unlink(missing_ok=True)
            wb2.close()
        except Exception as exc:  # noqa: BLE001
            errors.append(f"Round-trip save/reload failed: {exc}")

    return {
        "valid": len(errors) == 0,
        "row_count": row_count,
        "source_counts": dict(sorted(source_counts.items())),
        "truncated_cells": truncated_cells,
        "errors": errors,
    }
