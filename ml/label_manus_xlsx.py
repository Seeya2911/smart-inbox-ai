"""Apply the authoritative LLM teacher to a Manus-ready XLSX corpus.

The workbook intentionally contains only subject/body/source. The local manifest
provides stable source identity and provenance for each workbook row. The LLM
produces the production intent + priority labels; the deterministic rule engine is
run independently and retained only as a disagreement/provenance signal.

This tool deliberately has no human-review fields and does not use an LLM label as
"ground truth". Its output is a pseudo-labeled training candidate that must remain
separate from the protected evaluation set.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Set

from openpyxl import load_workbook

from ml.llm_labeler import LLMClient, OpenAILLMClient, label_example
from ml.schema import CanonicalEmailExample, format_namespaced_id

EXPECTED_HEADERS = ["subject", "body", "source"]


def _load_manifest(path: Path) -> Dict[int, Dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    row_map = payload.get("row_map")
    if not isinstance(row_map, list):
        raise ValueError("Manifest is missing row_map; regenerate the Manus corpus with the current exporter")

    result: Dict[int, Dict[str, Any]] = {}
    for item in row_map:
        if not isinstance(item, dict) or "output_row" not in item:
            raise ValueError("Manifest contains an invalid row_map entry")
        row_number = int(item["output_row"])
        if row_number in result:
            raise ValueError(f"Manifest contains duplicate output_row {row_number}")
        result[row_number] = item
    return result


def _placeholder_example(row_number: int, row: List[Any], metadata: Dict[str, Any]) -> CanonicalEmailExample:
    subject = "" if row[0] is None else str(row[0]).strip()
    body = "" if row[1] is None else str(row[1]).strip()
    source = str(metadata.get("source", row[2] if row[2] is not None else "")).strip()
    if not body:
        raise ValueError(f"Workbook row {row_number} has empty body")
    source_id = str(metadata.get("id", "")).strip()
    if not source_id:
        source_id = f"row-{row_number}"
    raw_id = format_namespaced_id(source, source_id)
    return CanonicalEmailExample(
        id=raw_id,
        subject=subject,
        body=body,
        intent="other",
        priority="low",
        source=source,
        source_example_id=str(metadata.get("source_example_id", source_id)),
        source_split=str(metadata.get("source_split", "unspecified")),
        label_source="rules",
        language=str(metadata.get("language", "en")).lower().strip() or "en",
        source_group_id=str(metadata.get("source_group_id", "")).strip(),
        is_synthetic=bool(metadata.get("is_synthetic", False)),
        provenance=str(metadata.get("provenance", source)).strip(),
    )


def load_workbook_examples(xlsx_path: Path, manifest_path: Path) -> List[CanonicalEmailExample]:
    """Load workbook rows and deterministically reattach manifest provenance."""
    mapping = _load_manifest(manifest_path)
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = [sheet.cell(row=1, column=i).value for i in range(1, 4)]
        if headers != EXPECTED_HEADERS:
            raise ValueError(f"Unexpected XLSX headers: {headers!r}; expected {EXPECTED_HEADERS!r}")

        examples: List[CanonicalEmailExample] = []
        seen_ids: Set[str] = set()
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, max_col=3, values_only=True), start=2):
            if all(value is None or not str(value).strip() for value in values):
                continue
            metadata = mapping.get(row_number)
            if metadata is None:
                raise ValueError(f"Workbook row {row_number} has no manifest mapping")
            example = _placeholder_example(row_number, list(values), metadata)
            if example.id in seen_ids:
                raise ValueError(f"Duplicate canonical example id in manifest: {example.id}")
            seen_ids.add(example.id)
            examples.append(example)

        expected_rows = len(mapping)
        if len(examples) != expected_rows:
            raise ValueError(f"Manifest maps {expected_rows} rows but workbook contains {len(examples)} usable rows")
        return examples
    finally:
        workbook.close()


def _existing_ids(path: Path) -> Set[str]:
    if not path.is_file():
        return set()
    ids: Set[str] = set()
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, 1):
            if not line.strip():
                continue
            payload = json.loads(line)
            example_id = str(payload.get("id", "")).strip()
            if not example_id:
                raise ValueError(f"Resume output row {line_number} is missing id")
            ids.add(example_id)
    return ids


def label_workbook(
    xlsx_path: Path,
    manifest_path: Path,
    output_path: Path,
    *,
    llm_client: LLMClient,
    resume: bool = False,
) -> Dict[str, Any]:
    """Label every workbook row with the LLM, optionally resuming a partial output."""
    examples = load_workbook_examples(xlsx_path, manifest_path)
    completed = _existing_ids(output_path) if resume else set()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    mode = "a" if resume else "w"
    labeled_count = 0
    disagreements = 0
    with output_path.open(mode, encoding="utf-8") as handle:
        for example in examples:
            if example.id in completed:
                continue
            labeled = label_example(example, llm_client)
            handle.write(json.dumps(labeled.to_dict(), ensure_ascii=False) + "\n")
            handle.flush()
            labeled_count += 1
            if labeled.llm_rule_agreement is False:
                disagreements += 1

    return {
        "status": "success",
        "input_rows": len(examples),
        "newly_labeled_rows": labeled_count,
        "already_labeled_rows": len(completed),
        "llm_rule_disagreements_new": disagreements,
        "output": str(output_path),
        "model": getattr(llm_client, "model", "custom"),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="LLM-label a Manus XLSX corpus for Smart Inbox AI")
    parser.add_argument("--input", required=True, type=Path, help="Manus XLSX with subject/body/source")
    parser.add_argument("--manifest", required=True, type=Path, help="Manifest generated alongside the XLSX")
    parser.add_argument("--output", required=True, type=Path, help="Canonical labeled JSONL output")
    parser.add_argument("--model", default=None, help="LLM model; defaults to SMART_INBOX_LLM_MODEL")
    parser.add_argument("--resume", action="store_true", help="Skip IDs already present in the output JSONL")
    args = parser.parse_args()

    client = OpenAILLMClient(model=args.model)
    summary = label_workbook(
        args.input,
        args.manifest,
        args.output,
        llm_client=client,
        resume=args.resume,
    )
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
