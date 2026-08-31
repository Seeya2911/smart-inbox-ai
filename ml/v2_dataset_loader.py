"""Dataset loader and auditor for smart_inbox_ai_dataset_v2.xlsx.

Loads the v2 dataset into CanonicalEmailExample instances, performs a full
pre-training audit (row count, column structure, label distributions,
intent×priority cross-tab, synthetic statistics, duplicate detection), and
returns only validated examples.

LEAKAGE SAFETY CONTRACT
-----------------------
Model features are constructed from subject + body ONLY.

The following columns are stored as provenance metadata in
CanonicalEmailExample but MUST NOT appear in the vectorized text
presented to any classifier:

    intent_reason   / priority_reason
    label_confidence
    source / label_source / is_synthetic
    index / id

The ``full_text`` property on CanonicalEmailExample correctly assembles:

    Subject: <subject>
    Body:
    <body>

This module never touches intent_reason, priority_reason, or
label_confidence when building model inputs.
"""
from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from ml.schema import (
    ALLOWED_INTENTS,
    ALLOWED_PRIORITIES,
    CanonicalEmailExample,
    format_namespaced_id,
)

# ---------------------------------------------------------------------------
# Column name aliases
# ---------------------------------------------------------------------------
# The loader tries each alias in order and uses the first one found.
_SUBJECT_COLS = ("subject", "Subject", "SUBJECT", "email_subject")
_BODY_COLS = ("body", "Body", "BODY", "email_body", "content", "text")
_INTENT_COLS = ("intent", "Intent", "INTENT", "canonical_intent", "label_intent")
_PRIORITY_COLS = ("priority", "Priority", "PRIORITY", "label_priority")
_INTENT_REASON_COLS = ("intent_reason", "IntentReason", "intent_reasoning", "reason_intent")
_PRIORITY_REASON_COLS = ("priority_reason", "PriorityReason", "priority_reasoning", "reason_priority")
_CONFIDENCE_COLS = ("label_confidence", "confidence", "LabelConfidence", "label_score")
_SOURCE_COLS = ("source", "Source", "SOURCE", "data_source", "dataset")
_LABEL_SOURCE_COLS = ("label_source", "LabelSource", "labeler", "annotator")
_IS_SYNTHETIC_COLS = ("is_synthetic", "IsSynthetic", "synthetic", "is_generated")
_ID_COLS = ("id", "ID", "Id", "email_id", "row_id", "index")

# Near-duplicate check settings
_ND_JACCARD_THRESHOLD = 0.85
_ND_MIN_LEN = 20


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _first_col(headers: List[str], aliases: Tuple[str, ...]) -> Optional[str]:
    """Return first alias present in headers, or None."""
    for alias in aliases:
        if alias in headers:
            return alias
    return None


def _normalize(text: str) -> str:
    """Lowercase, NFKC-normalize, collapse whitespace."""
    text = unicodedata.normalize("NFKC", text).lower()
    return re.sub(r"\s+", " ", text).strip()


def _md5(text: str) -> str:
    return hashlib.md5(text.encode("utf-8")).hexdigest()


def _word_jaccard(a: str, b: str) -> float:
    wa, wb = set(a.split()), set(b.split())
    if not wa and not wb:
        return 1.0
    if not wa or not wb:
        return 0.0
    return len(wa & wb) / len(wa | wb)


def _char_jaccard(a: str, b: str, n: int = 4) -> float:
    if a == b:
        return 1.0
    if len(a) < n or len(b) < n:
        return 1.0 if a == b else 0.0
    ng_a = {a[i: i + n] for i in range(len(a) - n + 1)}
    ng_b = {b[i: i + n] for i in range(len(b) - n + 1)}
    union = ng_a | ng_b
    return len(ng_a & ng_b) / len(union) if union else 0.0


def _to_bool(val: Any) -> bool:
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.strip().lower() in {"true", "1", "yes", "y"}
    return False


def _safe_float(val: Any, default: float = 1.0) -> float:
    try:
        f = float(val)
        return max(0.0, min(1.0, f))
    except (TypeError, ValueError):
        return default


# ---------------------------------------------------------------------------
# XLSX Loading
# ---------------------------------------------------------------------------

def load_xlsx_sheet(path: Path) -> Tuple[List[str], List[List[Any]]]:
    """Load the first (or 'emails') sheet from an xlsx file.

    Returns (headers, data_rows) where data_rows excludes the header.
    Raises FileNotFoundError or ValueError on fatal issues.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl is required: pip install openpyxl") from exc

    if not path.exists():
        raise FileNotFoundError(f"Dataset file not found: {path}")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    # Prefer 'emails' sheet; fall back to first sheet
    sheet_name = "emails" if "emails" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_header = list(next(rows_iter))
    except StopIteration:
        wb.close()
        raise ValueError("Worksheet is completely empty (no header row)")

    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(raw_header)]
    data_rows: List[List[Any]] = [list(r) for r in rows_iter]
    wb.close()

    return headers, data_rows


# ---------------------------------------------------------------------------
# Row → CanonicalEmailExample conversion
# ---------------------------------------------------------------------------

class _ColumnMap:
    """Maps semantic field names to their actual column indices."""

    def __init__(self, headers: List[str]) -> None:
        self.headers = headers
        self._idx: Dict[str, int] = {h: i for i, h in enumerate(headers)}

    def col_idx(self, aliases: Tuple[str, ...]) -> Optional[int]:
        for alias in aliases:
            if alias in self._idx:
                return self._idx[alias]
        return None

    def get(self, row: List[Any], aliases: Tuple[str, ...], default: Any = None) -> Any:
        idx = self.col_idx(aliases)
        if idx is None:
            return default
        val = row[idx] if idx < len(row) else default
        return val


def _row_to_example(
    row: List[Any],
    row_number: int,
    col_map: "_ColumnMap",
) -> Tuple[Optional[CanonicalEmailExample], Optional[str]]:
    """Convert one xlsx row to CanonicalEmailExample.

    Returns (example, None) on success, (None, rejection_reason) on failure.
    Intent, priority, source, label_source, and label_confidence are preserved
    as CanonicalEmailExample metadata fields.  They are NOT included in
    full_text / model features.
    """
    # --- Raw field extraction ---
    raw_subject = str(col_map.get(row, _SUBJECT_COLS, "") or "").strip()
    raw_body = str(col_map.get(row, _BODY_COLS, "") or "").strip()
    raw_intent = str(col_map.get(row, _INTENT_COLS, "") or "").strip().lower()
    raw_priority = str(col_map.get(row, _PRIORITY_COLS, "") or "").strip().lower()

    # Metadata (not used as features)
    raw_intent_reason = str(col_map.get(row, _INTENT_REASON_COLS, "") or "").strip()
    raw_priority_reason = str(col_map.get(row, _PRIORITY_REASON_COLS, "") or "").strip()
    raw_confidence = col_map.get(row, _CONFIDENCE_COLS, 1.0)
    raw_source = str(col_map.get(row, _SOURCE_COLS, "synthetic") or "synthetic").strip()
    raw_label_source = str(col_map.get(row, _LABEL_SOURCE_COLS, "llm") or "llm").strip().lower()
    raw_is_synthetic = col_map.get(row, _IS_SYNTHETIC_COLS, None)
    raw_id = str(col_map.get(row, _ID_COLS, "") or "").strip()

    # --- Hard rejections ---
    if not raw_body:
        return None, f"Row {row_number}: empty body"
    if not raw_intent:
        return None, f"Row {row_number}: missing intent"
    if not raw_priority:
        return None, f"Row {row_number}: missing priority"
    if raw_intent not in ALLOWED_INTENTS:
        return None, f"Row {row_number}: invalid intent {raw_intent!r}"
    if raw_priority not in ALLOWED_PRIORITIES:
        return None, f"Row {row_number}: invalid priority {raw_priority!r}"

    # --- ID normalization ---
    if not raw_id:
        raw_id = str(row_number)
    namespaced_id = format_namespaced_id(raw_source, raw_id)

    # --- Infer is_synthetic ---
    if raw_is_synthetic is not None:
        is_synthetic = _to_bool(raw_is_synthetic)
    else:
        is_synthetic = "synthetic" in raw_source.lower()

    # --- Label source normalization ---
    from ml.schema import ALLOWED_LABEL_SOURCES
    if raw_label_source not in ALLOWED_LABEL_SOURCES:
        raw_label_source = "llm"

    confidence = _safe_float(raw_confidence, 1.0)

    # --- priority_reasons: store reason as list for provenance ---
    priority_reasons: List[str] = []
    if raw_priority_reason:
        priority_reasons.append(raw_priority_reason)

    try:
        ex = CanonicalEmailExample(
            id=namespaced_id,
            subject=raw_subject,
            body=raw_body,
            intent=raw_intent,
            priority=raw_priority,
            priority_reasons=priority_reasons,
            source=raw_source,
            label_source=raw_label_source,
            label_confidence=confidence,
            rule_score=0.0,
            language="en",
            source_group_id="",
            is_synthetic=is_synthetic,
            provenance=f"v2_dataset_row_{row_number}",
            llm_intent_reason=raw_intent_reason,
            llm_priority_reason=raw_priority_reason,
        )
    except ValueError as exc:
        return None, f"Row {row_number}: schema validation failed: {exc}"

    return ex, None


# ---------------------------------------------------------------------------
# Audit helpers
# ---------------------------------------------------------------------------

def _distribution(values: Sequence[str]) -> Dict[str, int]:
    return dict(sorted(Counter(values).items()))


def _cross_tab(
    intents: Sequence[str],
    priorities: Sequence[str],
    all_intents: Sequence[str],
    all_priorities: Sequence[str],
) -> Dict[str, Dict[str, int]]:
    """Return intent×priority count table, including zero-count combinations."""
    tab: Dict[str, Dict[str, int]] = {
        intent: {p: 0 for p in sorted(all_priorities)} for intent in sorted(all_intents)
    }
    for intent, priority in zip(intents, priorities):
        if intent in tab and priority in tab[intent]:
            tab[intent][priority] += 1
    return tab


def _near_duplicate_report(
    examples: List[CanonicalEmailExample],
    threshold: float = _ND_JACCARD_THRESHOLD,
) -> Dict[str, Any]:
    """Report near-duplicate pairs without removing anything."""
    normalized = [_normalize(ex.full_text) for ex in examples]
    near_dup_pairs = 0
    sampled: List[Tuple[str, str]] = []

    for i in range(len(normalized)):
        a = normalized[i]
        len_a = len(a)
        if len_a < _ND_MIN_LEN:
            continue
        for j in range(i + 1, len(normalized)):
            b = normalized[j]
            len_b = len(b)
            if len_b < _ND_MIN_LEN:
                continue
            if min(len_a, len_b) / max(len_a, len_b) < 0.60:
                continue
            w_sim = _word_jaccard(a, b)
            if w_sim < 0.35:
                continue
            sim = _char_jaccard(a, b)
            if sim >= threshold:
                near_dup_pairs += 1
                if len(sampled) < 5:
                    sampled.append((examples[i].id, examples[j].id))

    return {"near_duplicate_pairs": near_dup_pairs, "sample_pairs": sampled}


def _synthetic_report(examples: List[CanonicalEmailExample]) -> Dict[str, Any]:
    """Report synthetic rows by intent and priority."""
    syn = [ex for ex in examples if ex.is_synthetic]
    return {
        "total_synthetic": len(syn),
        "total_real": len(examples) - len(syn),
        "synthetic_by_intent": _distribution([ex.intent for ex in syn]),
        "synthetic_by_priority": _distribution([ex.priority for ex in syn]),
    }


# ---------------------------------------------------------------------------
# Main public API
# ---------------------------------------------------------------------------

def load_v2_dataset(
    path: Path,
    print_audit: bool = True,
) -> Tuple[List[CanonicalEmailExample], Dict[str, Any]]:
    """Load and audit smart_inbox_ai_dataset_v2.xlsx.

    Parameters
    ----------
    path:
        Path to the xlsx file (read-only; never modified).
    print_audit:
        If True, print a human-readable audit report to stdout.

    Returns
    -------
    (examples, audit_report)
        examples     — validated CanonicalEmailExample list (duplicates removed).
        audit_report — full audit dict suitable for JSON serialization.

    Raises
    ------
    FileNotFoundError, ValueError
        On fatal load issues.

    LEAKAGE NOTE
    ------------
    This function constructs CanonicalEmailExample objects where:
        full_text = "Subject: <subject>\\nBody:\\n<body>"
    The fields intent_reason, priority_reason, label_confidence, source,
    label_source, is_synthetic, and id are stored as metadata fields only.
    They do NOT appear in full_text and MUST NOT be passed to any vectorizer.
    """
    headers, data_rows = load_xlsx_sheet(path)
    col_map = _ColumnMap(headers)

    # ---- Report actual columns found ----
    missing_warnings: List[str] = []
    actual_cols = {
        "subject": _first_col(headers, _SUBJECT_COLS),
        "body": _first_col(headers, _BODY_COLS),
        "intent": _first_col(headers, _INTENT_COLS),
        "priority": _first_col(headers, _PRIORITY_COLS),
        "intent_reason": _first_col(headers, _INTENT_REASON_COLS),
        "priority_reason": _first_col(headers, _PRIORITY_REASON_COLS),
        "label_confidence": _first_col(headers, _CONFIDENCE_COLS),
        "source": _first_col(headers, _SOURCE_COLS),
        "label_source": _first_col(headers, _LABEL_SOURCE_COLS),
        "is_synthetic": _first_col(headers, _IS_SYNTHETIC_COLS),
        "id": _first_col(headers, _ID_COLS),
    }
    required = ("subject", "body", "intent", "priority")
    for field in required:
        if actual_cols[field] is None:
            missing_warnings.append(f"Required column '{field}' not found in headers: {headers}")

    if missing_warnings:
        raise ValueError(
            "Dataset is missing required columns:\n" + "\n".join(missing_warnings)
        )

    # ---- Parse rows ----
    valid_examples: List[CanonicalEmailExample] = []
    rejections: List[Dict[str, str]] = []

    for row_idx, row in enumerate(data_rows, start=2):
        ex, reason = _row_to_example(row, row_idx, col_map)
        if ex is not None:
            valid_examples.append(ex)
        else:
            rejections.append({"row": row_idx, "reason": reason or "unknown"})

    # ---- Duplicate ID check ----
    id_counter: Counter[str] = Counter(ex.id for ex in valid_examples)
    duplicate_ids = {eid: cnt for eid, cnt in id_counter.items() if cnt > 1}

    # ---- Normalized-text duplicate check ----
    seen_hashes: Dict[str, str] = {}
    exact_dup_ids: List[str] = []
    deduped: List[CanonicalEmailExample] = []
    for ex in valid_examples:
        norm = _normalize(ex.full_text)
        h = _md5(norm)
        if h in seen_hashes:
            exact_dup_ids.append(ex.id)
        else:
            seen_hashes[h] = ex.id
            deduped.append(ex)

    # ---- Distributions ----
    intents = [ex.intent for ex in deduped]
    priorities = [ex.priority for ex in deduped]
    intent_dist = _distribution(intents)
    priority_dist = _distribution(priorities)
    joint_dist_raw = _cross_tab(
        intents, priorities,
        all_intents=sorted(ALLOWED_INTENTS),
        all_priorities=sorted(ALLOWED_PRIORITIES),
    )

    # ---- Near-duplicate report ----
    # Only run on deduped set; cap at 1000 examples for speed
    nd_sample = deduped[:1000]
    near_dup_report = _near_duplicate_report(nd_sample, threshold=_ND_JACCARD_THRESHOLD)

    # ---- Synthetic report ----
    syn_report = _synthetic_report(deduped)

    # ---- Low-confidence report ----
    low_conf = [ex for ex in deduped if ex.label_confidence < 0.7]
    low_conf_count = len(low_conf)

    # ---- Source distribution ----
    source_dist = _distribution([ex.source for ex in deduped])
    label_source_dist = _distribution([ex.label_source for ex in deduped])

    # ---- Class imbalance warnings ----
    imbalance_warnings: List[str] = []
    if intent_dist:
        max_intent = max(intent_dist.values())
        for cls, cnt in intent_dist.items():
            if cnt < 10:
                imbalance_warnings.append(
                    f"Intent '{cls}' has only {cnt} example(s) — minority class warning"
                )
            elif max_intent / cnt > 10:
                imbalance_warnings.append(
                    f"Intent '{cls}' imbalance ratio {max_intent/cnt:.1f}:1"
                )

    audit: Dict[str, Any] = {
        "xlsx_path": str(path),
        "xlsx_size_bytes": path.stat().st_size if path.exists() else 0,
        "sheet_headers": headers,
        "actual_column_mapping": actual_cols,
        "total_rows_in_sheet": len(data_rows),
        "valid_rows": len(deduped),
        "rejected_rows": len(rejections),
        "rejections": rejections[:50],  # cap for readability
        "exact_duplicate_normalized_texts": len(exact_dup_ids),
        "duplicate_ids": duplicate_ids,
        "near_duplicate_report": near_dup_report,
        "intent_distribution": intent_dist,
        "priority_distribution": priority_dist,
        "intent_x_priority_distribution": joint_dist_raw,
        "source_distribution": source_dist,
        "label_source_distribution": label_source_dist,
        "synthetic_report": syn_report,
        "low_confidence_count": low_conf_count,
        "imbalance_warnings": imbalance_warnings,
    }

    if print_audit:
        _print_audit(audit)

    return deduped, audit


def _print_audit(audit: Dict[str, Any]) -> None:
    """Print human-readable audit report."""
    sep = "=" * 62
    print(f"\n{sep}")
    print("  SMART INBOX AI v2 DATASET AUDIT")
    print(sep)
    print(f"  File  : {audit['xlsx_path']}")
    print(f"  Size  : {audit['xlsx_size_bytes'] / 1024:.1f} KB")
    print(f"  Headers ({len(audit['sheet_headers'])}): {audit['sheet_headers']}")
    print()
    print(f"  Total rows in sheet : {audit['total_rows_in_sheet']}")
    print(f"  Valid rows          : {audit['valid_rows']}")
    print(f"  Rejected rows       : {audit['rejected_rows']}")
    print(f"  Exact norm-dups rem : {audit['exact_duplicate_normalized_texts']}")
    print(f"  Duplicate IDs       : {len(audit['duplicate_ids'])}")
    print()

    print("  INTENT DISTRIBUTION")
    for intent, cnt in sorted(audit["intent_distribution"].items()):
        bar = "#" * min(40, int(cnt * 40 / max(audit["intent_distribution"].values(), default=1)))
        print(f"    {intent:<20} {cnt:>4}  {bar}")
    print()

    print("  PRIORITY DISTRIBUTION")
    for pri, cnt in sorted(audit["priority_distribution"].items()):
        bar = "#" * min(40, int(cnt * 40 / max(audit["priority_distribution"].values(), default=1)))
        print(f"    {pri:<20} {cnt:>4}  {bar}")
    print()

    print("  INTENT × PRIORITY CROSS-TAB")
    priorities = sorted(ALLOWED_PRIORITIES)
    header_row = f"  {'intent':<22}" + "".join(f"{p:>8}" for p in priorities)
    print(header_row)
    print("  " + "-" * (22 + 8 * len(priorities)))
    for intent in sorted(ALLOWED_INTENTS):
        row_data = audit["intent_x_priority_distribution"].get(intent, {})
        row_str = f"  {intent:<22}" + "".join(f"{row_data.get(p, 0):>8}" for p in priorities)
        print(row_str)
    print()

    print("  SYNTHETIC REPORT")
    syn = audit["synthetic_report"]
    print(f"    Total synthetic : {syn['total_synthetic']}")
    print(f"    Total real      : {syn['total_real']}")
    print(f"    Synthetic by intent  : {syn['synthetic_by_intent']}")
    print(f"    Synthetic by priority: {syn['synthetic_by_priority']}")
    print()

    nd = audit["near_duplicate_report"]
    print(f"  NEAR-DUPLICATE PAIRS (Jaccard>=0.85): {nd['near_duplicate_pairs']}")
    if nd["sample_pairs"]:
        print(f"    Sample pairs: {nd['sample_pairs'][:3]}")
    print()

    print(f"  LOW-CONFIDENCE ROWS (< 0.7): {audit['low_confidence_count']}")
    print()

    if audit["imbalance_warnings"]:
        print("  [WARN] IMBALANCE WARNINGS")
        for w in audit["imbalance_warnings"]:
            print(f"    - {w}")
        print()

    if audit["rejections"]:
        print("  [FAIL] REJECTIONS (first 10)")
        for r in audit["rejections"][:10]:
            print(f"    Row {r['row']}: {r['reason']}")
        print()

    print(f"  SOURCE DISTRIBUTION: {audit['source_distribution']}")
    print(f"  LABEL SOURCE      : {audit['label_source_distribution']}")
    print(sep + "\n")


# ---------------------------------------------------------------------------
# JSONL export helper
# ---------------------------------------------------------------------------

def export_to_jsonl(examples: List[CanonicalEmailExample], output_path: Path) -> None:
    """Write examples to JSONL for pipeline consumption."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as fh:
        for ex in examples:
            fh.write(json.dumps(ex.to_dict(), ensure_ascii=False) + "\n")


def load_from_jsonl(path: Path) -> List[CanonicalEmailExample]:
    """Load CanonicalEmailExample list from a previously exported JSONL."""
    examples: List[CanonicalEmailExample] = []
    with path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                examples.append(CanonicalEmailExample.from_dict(json.loads(line)))
    return examples
