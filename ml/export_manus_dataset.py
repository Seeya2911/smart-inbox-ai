"""Build a deterministic Manus-ready XLSX from the canonical raw email pool.

The exporter intentionally contains NO Smart Inbox labels. It selects clean,
deduplicated email text and preserves provenance in a local manifest so the
LLM-labeled results can be joined back to the canonical records deterministically.

The output workbook contains exactly three labeling columns: subject, body, source.
The manifest contains row identity/provenance metadata and is not part of the
external labeling input.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from ml.xlsx_utils import EXCEL_MAX_CELL_CHARS, prepare_cell, validate_workbook


def _norm(text: str) -> str:
    text = text.lower().replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _content_key(row: dict[str, Any]) -> str:
    subject = _norm(str(row.get("subject", "")))
    body = _norm(str(row.get("body", "")))
    return hashlib.sha256(f"{subject}\n{body}".encode("utf-8")).hexdigest()


def load_rows(paths: list[Path]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """Load, deduplicate, and return rows from canonical JSONL files."""
    rows: list[dict[str, Any]] = []
    stats: dict[str, int] = {
        "input_rows": 0,
        "empty_removed": 0,
        "exact_duplicates_removed": 0,
    }
    seen: set[str] = set()
    for path in paths:
        with path.open("r", encoding="utf-8") as handle:
            for line in handle:
                if not line.strip():
                    continue
                stats["input_rows"] += 1
                row = json.loads(line)
                subject = str(row.get("subject", "")).strip()
                body = str(row.get("body", "")).strip()
                source = str(row.get("source", "")).strip()
                if not subject and not body:
                    stats["empty_removed"] += 1
                    continue
                if not source:
                    raise ValueError("Every row must have a non-empty source")
                row["subject"] = subject
                row["body"] = body
                row["source"] = source
                key = _content_key(row)
                if key in seen:
                    stats["exact_duplicates_removed"] += 1
                    continue
                seen.add(key)
                rows.append(row)
    rows.sort(
        key=lambda r: (str(r["source"]), _norm(str(r["subject"])), str(r.get("id", "")))
    )
    return rows, stats


def export_xlsx(rows: list[dict[str, Any]], output: Path) -> dict[str, int]:
    """Write *rows* to a Manus-ready XLSX at *output*."""
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "emails"

    for col_idx, col_name in enumerate(["subject", "body", "source"], start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.data_type = "s"

    truncated_rows = 0
    for row_idx, row in enumerate(rows, start=2):
        subject_cell, sub_trunc = prepare_cell(row["subject"])
        body_cell, body_trunc = prepare_cell(row["body"])
        source_cell, _src_trunc = prepare_cell(row["source"])
        if sub_trunc or body_trunc:
            truncated_rows += 1

        for col_idx, val in enumerate([subject_cell, body_cell, source_cell], start=1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.data_type = "s"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:C{len(rows) + 1}"
    sheet.column_dimensions["A"].width = 45
    sheet.column_dimensions["B"].width = 100
    sheet.column_dimensions["C"].width = 20
    workbook.save(output)
    return {"truncated_rows": truncated_rows}


def write_manifest(
    rows: list[dict[str, Any]],
    stats: dict[str, int],
    output: Path,
) -> None:
    """Write counts plus deterministic workbook-row-to-source mapping."""
    source_counts: dict[str, int] = {}
    row_map: list[dict[str, Any]] = []
    for output_row, row in enumerate(rows, start=2):
        source = row["source"]
        source_counts[source] = source_counts.get(source, 0) + 1
        row_map.append(
            {
                "output_row": output_row,
                "id": str(row.get("id", "")),
                "source_example_id": str(row.get("source_example_id", row.get("id", ""))),
                "source": source,
                "source_dataset": str(row.get("source_dataset", "")),
                "source_split": str(row.get("source_split", "unspecified")),
                "source_group_id": str(row.get("source_group_id", row.get("group_id", ""))),
                "language": str(row.get("language", "en")),
                "is_synthetic": bool(row.get("is_synthetic", False)),
                "provenance": str(row.get("provenance", source)),
            }
        )
    payload = {
        **stats,
        "output_rows": len(rows),
        "source_counts": dict(sorted(source_counts.items())),
        "excel_max_cell_chars": EXCEL_MAX_CELL_CHARS,
        "row_map": row_map,
    }
    output.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Export clean canonical emails for external labeling")
    parser.add_argument("inputs", nargs="+", type=Path, help="canonical JSONL files")
    parser.add_argument("--output", type=Path, required=True, help="labeling XLSX")
    parser.add_argument("--manifest", type=Path, default=None)
    args = parser.parse_args()

    rows, stats = load_rows(args.inputs)
    if not rows:
        raise SystemExit("No usable email rows found")

    export_stats = export_xlsx(rows, args.output)
    stats.update(export_stats)

    manifest = args.manifest or args.output.with_suffix(".manifest.json")
    write_manifest(rows, stats, manifest)

    print("Validating generated workbook …", flush=True)
    result = validate_workbook(args.output)
    if not result["valid"]:
        print("XLSX VALIDATION FAILED:", file=sys.stderr)
        for err in result["errors"]:
            print(f"  ERROR: {err}", file=sys.stderr)
        raise SystemExit(
            f"Workbook validation failed with {len(result['errors'])} error(s). "
            "The XLSX will NOT be uploaded."
        )

    summary = {
        **stats,
        "output_rows": len(rows),
        "manifest": str(manifest),
        "validation": {
            "valid": result["valid"],
            "row_count": result["row_count"],
            "source_counts": result["source_counts"],
            "truncated_cells": result["truncated_cells"],
        },
    }
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
